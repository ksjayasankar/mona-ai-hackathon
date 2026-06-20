"""RosterSink tests — fully offline (xlsx write-back; no Google creds)."""
from datetime import datetime

import openpyxl

from services.roster_sink import SyncResult, XlsxSink, get_sink


def _rows():
    return [
        {"employee_id": "HOSP-1059", "name": "Felix Haddad", "role": "Registered Nurse",
         "department": "ICU", "Fri 06/19": "O", "Sat 06/20": "O"},
        {"employee_id": "HOSP-2007", "name": "Anya Lindgren", "role": "Registered Nurse",
         "department": "ICU", "Fri 06/19": "O", "Sat 06/20": "O"},
    ]


def test_xlsx_sink_push_then_record_fill(tmp_path):
    out = tmp_path / "updated.xlsx"
    sink = XlsxSink(out_path=out)
    r = sink.push_roster(_rows(), ["Fri 06/19", "Sat 06/20"])
    assert r.ok and r.target == "xlsx" and out.exists()
    res = sink.record_fill(employee_id="HOSP-2007", name="Anya Lindgren",
                           day_label="Sat 06/20", code="N", when=datetime(2026, 6, 20, 18, 45))
    assert res.ok and res.target == "xlsx"
    wb = openpyxl.load_workbook(out)
    ws = wb["Roster"]
    header = [c.value for c in ws[1]]
    sat = header.index("Sat 06/20")
    anya = next(row for row in ws.iter_rows(min_row=2, values_only=True) if row[0] == "HOSP-2007")
    assert anya[sat] == "N"
    assert "Updates" in wb.sheetnames
    updates = list(wb["Updates"].iter_rows(min_row=2, values_only=True))
    assert any(u[0] == "HOSP-2007" and u[2] == "Sat 06/20" and u[3] == "N" for u in updates)


def test_get_sink_defaults_to_xlsx_without_creds(monkeypatch):
    monkeypatch.delenv("GOOGLE_SHEETS_ID", raising=False)
    monkeypatch.delenv("GOOGLE_SHEETS_CREDENTIALS_JSON", raising=False)
    assert isinstance(get_sink(), XlsxSink)
    assert isinstance(SyncResult("xlsx", True, None), SyncResult)

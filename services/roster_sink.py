"""Pluggable roster sink: where a shift-fill lands in the 'real world'.

XlsxSink (default, zero-creds) writes data/hospital_schedule_updated.xlsx with the flipped
cell highlighted + an Updates audit sheet. GoogleSheetsSink (roster_sink_google) flips the
same cell in a live Google Sheet when GOOGLE_SHEETS_ID + a service-account key are configured.
Both are best-effort: callers wrap them so a sync failure never breaks a fill."""
from __future__ import annotations

import logging
import os
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Protocol

import openpyxl
from openpyxl.styles import PatternFill

from core import config

log = logging.getLogger("roster_sink")
_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


@dataclass
class SyncResult:
    target: str            # "google_sheets" | "xlsx" | "none"
    ok: bool
    link: str | None
    detail: str | None = None


class RosterSink(Protocol):
    def push_roster(self, rows: list[dict], day_cols: list[str]) -> SyncResult: ...
    def record_fill(self, *, employee_id: str, name: str, day_label: str,
                    code: str, when: datetime) -> SyncResult: ...


class XlsxSink:
    """Writes an updated .xlsx the customer can open in Excel / drag into Sheets."""

    def __init__(self, out_path: Path | str | None = None, source_path: Path | str | None = None):
        self.out_path = Path(out_path) if out_path else config.DATA_OUT / "hospital_schedule_updated.xlsx"
        self.source_path = Path(source_path) if source_path else config.PATHS["schedule"]

    def push_roster(self, rows: list[dict], day_cols: list[str]) -> SyncResult:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Roster"
        ws.append(["employee_id", "name", "role", "department", *day_cols])
        for r in rows:
            ws.append([r.get("employee_id"), r.get("name"), r.get("role"),
                       r.get("department"), *[r.get(c, "") for c in day_cols]])
        up = wb.create_sheet("Updates")
        up.append(["employee_id", "name", "day", "code", "when"])
        self.out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.out_path)
        return SyncResult("xlsx", True, str(self.out_path))

    def record_fill(self, *, employee_id: str, name: str, day_label: str,
                    code: str, when: datetime) -> SyncResult:
        if not self.out_path.exists():
            # no roster pushed yet — seed a minimal sheet so the fill is still recordable
            self.push_roster([{"employee_id": employee_id, "name": name, day_label: "O"}], [day_label])
        wb = openpyxl.load_workbook(self.out_path)
        ws = wb["Roster"]
        header = [c.value for c in ws[1]]
        if day_label not in header:
            ws.cell(row=1, column=len(header) + 1, value=day_label)
            header.append(day_label)
        col = header.index(day_label) + 1
        for row in ws.iter_rows(min_row=2):
            if row[0].value == employee_id:
                cell = ws.cell(row=row[0].row, column=col, value=code)
                cell.fill = _GREEN
                break
        up = wb["Updates"] if "Updates" in wb.sheetnames else wb.create_sheet("Updates")
        up.append([employee_id, name, day_label, code, when.isoformat(timespec="minutes")])
        wb.save(self.out_path)
        return SyncResult("xlsx", True, str(self.out_path))


def get_sink() -> RosterSink:
    """Pick the configured sink: Apps Script webhook -> gspread service account -> xlsx."""
    if os.getenv("GOOGLE_APPS_SCRIPT_URL"):
        try:
            from services.roster_sink_apps_script import AppsScriptSink  # lazy
            return AppsScriptSink()
        except Exception as e:                       # bad config -> fall through
            log.warning("AppsScriptSink unavailable (%s); falling back", e)
    if os.getenv("GOOGLE_SHEETS_ID") and os.getenv("GOOGLE_SHEETS_CREDENTIALS_JSON"):
        try:
            from services.roster_sink_google import GoogleSheetsSink  # lazy: avoids gspread import by default
            return GoogleSheetsSink()
        except Exception as e:                       # missing dep / bad creds -> fall back
            log.warning("GoogleSheetsSink unavailable (%s); using xlsx", e)
    return XlsxSink()
